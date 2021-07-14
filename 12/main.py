import openpyxl

def main():

    wb = openpyxl.Workbook()
    #创建文件

    sheet1=wb.create_sheet('sh1')
    sheet2 =wb.create_sheet('sh2')
    #创建sheet

    wb.save('book.xlsx')
    # 保存
    wb.close()
    # 关闭

    print('__________________________________')


    workbook=openpyxl.load_workbook(r'work.xlsx')
    #载入excel

    sh=workbook['Sheet']
    #或者  sh=workbook.worksheets[0]  第x章就是第x-1个元素
    #载入sheet

    r=sh.max_row
    c=sh.max_column
    # sheet的最大行列 row行 column列

    # 读取数据  行列从1开始
    for m in range(1,r+1):
        for n in range (1,c+1):
            res = sh.cell(row=m, column=n)
            #获取格子
            value=str(res.value)
            # 获取值 类型未定
            print(value.rjust(10),end='\t'*3)

        print()

    shcp = workbook['Sheetcopy']
    shcp.cell(1, 2, "so do i")
    # 写入

    workbook.save('work.xlsx')


if __name__ == '__main__':
    main()

# 访问 https://www.jetbrains.com/help/pycharm/ 获取 PyCharm 帮助
